"""Tools del Excel Agent: exportación a workbooks personalizados.

Complementa al Builder: el dashboard vive en Looker (interactivo, gobernado),
y este agente produce el entregable offline — un .xlsx con formato profesional
(cabeceras, autofiltro, anchos, formatos numéricos, hoja por tile) — para los
flujos donde el usuario necesita 'el Excel': envíos a terceros, cierres, etc.

Los datos salen SIEMPRE de queries de la capa semántica (mismo principio
anti-alucinación: campos validados por el Catalog). La entrega es por URL
firmada de GCS (v4, expiración configurable): los bytes nunca pasan por el
texto del modelo.
"""
import datetime
import io
import json
import os
import re

import looker_sdk.sdk.api40.models as models40
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common.looker_client import get_sdk

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=11)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MAX_COL_WIDTH = 60
SIGNED_URL_HOURS = int(os.environ.get("EXCEL_URL_EXPIRY_HOURS", "24"))


# ---------------------------------------------------------------------------
# Construcción del workbook
# ---------------------------------------------------------------------------
def _safe_sheet_name(name: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "-", name)[:31] or "Datos"


def _write_sheet(wb: Workbook, sheet_name: str, rows: list[dict], title: str = ""):
    ws = wb.create_sheet(_safe_sheet_name(sheet_name))
    start = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, name="Calibri")
        ws.cell(row=2, column=1,
                value=f"Generado: {datetime.datetime.now():%Y-%m-%d %H:%M}").font = Font(
                    italic=True, size=9, color="808080", name="Calibri")
        start = 4
    if not rows:
        ws.cell(row=start, column=1, value="Sin datos para los filtros aplicados.")
        return ws

    headers = list(rows[0].keys())
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h.split(".")[-1].replace("_", " ").title())
        cell.fill, cell.font, cell.border = HEADER_FILL, HEADER_FONT, BORDER
        cell.alignment = Alignment(horizontal="center")

    widths = {c: len(str(ws.cell(row=start, column=c).value)) for c in range(1, len(headers) + 1)}
    for r, row in enumerate(rows, start + 1):
        for c, h in enumerate(headers, 1):
            v = row.get(h)
            cell = ws.cell(row=r, column=c, value=v)
            cell.font, cell.border = BODY_FONT, BORDER
            if isinstance(v, float):
                cell.number_format = "#,##0.00"
            elif isinstance(v, int):
                cell.number_format = "#,##0"
            widths[c] = min(MAX_COL_WIDTH, max(widths[c], len(str(v)) if v is not None else 0))

    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w + 3
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{start}:{last_col}{start + len(rows)}"
    ws.freeze_panes = ws.cell(row=start + 1, column=1)
    return ws


def _base_workbook(template_name: str = "") -> Workbook:
    """Workbook vacío o cargado desde un template corporativo en GCS
    (portada con branding, estilos, hojas fijas se preservan)."""
    if not template_name:
        return Workbook()
    from openpyxl import load_workbook
    from common import templates
    name = template_name if template_name.endswith(".xlsx") else f"{template_name}.xlsx"
    return load_workbook(io.BytesIO(templates.load_bytes("excel", name)))


def list_excel_templates() -> str:
    """Lista los templates de Excel corporativos disponibles."""
    from common import templates
    return json.dumps(templates.list_names("excel"))


def _finish_workbook(wb: Workbook) -> bytes:
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Entrega: GCS + URL firmada v4
# ---------------------------------------------------------------------------
def _upload_and_sign(data: bytes, filename: str) -> str:
    import google.auth
    from google.auth.transport import requests as ga_requests
    from google.cloud import storage

    bucket_name = os.environ["EXPORT_BUCKET"]
    creds, _ = google.auth.default()
    creds.refresh(ga_requests.Request())
    client = storage.Client(credentials=creds)
    blob = client.bucket(bucket_name).blob(f"exports/{filename}")
    blob.upload_from_string(
        data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # Firma v4 sin llave privada: usa signBlob vía IAM (requiere
    # roles/iam.serviceAccountTokenCreator de la SA sobre sí misma).
    return blob.generate_signed_url(
        version="v4",
        expiration=datetime.timedelta(hours=SIGNED_URL_HOURS),
        service_account_email=getattr(creds, "service_account_email", None),
        access_token=creds.token,
        method="GET",
    )


def _run_query_rows(model: str, explore: str, fields: list[str],
                    filters: dict | None, sorts: list[str] | None, limit: int) -> list[dict]:
    sdk = get_sdk()
    q = models40.WriteQuery(model=model, view=explore, fields=fields,
                            filters=filters or {}, sorts=sorts or [], limit=str(limit))
    raw = sdk.run_inline_query(result_format="json", body=q)
    return json.loads(raw) if isinstance(raw, str) else raw


# ---------------------------------------------------------------------------
# Tools
# ---------------------------------------------------------------------------
def export_query_to_excel(model: str, explore: str, fields: list[str],
                          filename: str, sheet_name: str = "Datos", title: str = "",
                          filters: dict | None = None, sorts: list[str] | None = None,
                          limit: int = 5000, template_name: str = "") -> str:
    """Exporta un query a un .xlsx con formato (cabeceras, autofiltro, anchos,
    formatos numéricos). fields: nombres exactos view.field validados por el
    Catalog Agent. Devuelve la URL firmada de descarga y el número de filas."""
    rows = _run_query_rows(model, explore, fields, filters, sorts, limit)
    wb = _base_workbook(template_name)
    _write_sheet(wb, sheet_name, rows, title=title or sheet_name)
    url = _upload_and_sign(_finish_workbook(wb), _xlsx_name(filename))
    return json.dumps({"download_url": url, "rows": len(rows),
                       "expires_in_hours": SIGNED_URL_HOURS})


def export_multi_sheet_excel(filename: str, sheets: list[dict],
                             template_name: str = "") -> str:
    """Exporta varias hojas en un solo workbook. Cada elemento de `sheets`:
    {"sheet_name": str, "title": str, "model": str, "explore": str,
     "fields": [..], "filters": {..} (opcional), "sorts": [..] (opcional),
     "limit": int (opcional)}."""
    wb = _base_workbook(template_name)
    total = 0
    for s in sheets:
        rows = _run_query_rows(s["model"], s["explore"], s["fields"],
                               s.get("filters"), s.get("sorts"), int(s.get("limit", 5000)))
        _write_sheet(wb, s["sheet_name"], rows, title=s.get("title", s["sheet_name"]))
        total += len(rows)
    url = _upload_and_sign(_finish_workbook(wb), _xlsx_name(filename))
    return json.dumps({"download_url": url, "sheets": len(sheets), "rows": total,
                       "expires_in_hours": SIGNED_URL_HOURS})


def export_dashboard_to_excel(dashboard_id: str, filename: str = "",
                              filters: dict | None = None,
                              template_name: str = "") -> str:
    """Exporta un dashboard existente a un workbook: una hoja por tile de datos,
    reutilizando los queries de los tiles. Ideal tras crear un dashboard con el
    Builder ('y mándamelo también en Excel')."""
    sdk = get_sdk()
    db = sdk.dashboard(dashboard_id=dashboard_id, fields="id,title,dashboard_elements")
    wb = _base_workbook(template_name)
    exported = []
    for el in (db.dashboard_elements or []):
        if el.type != "vis" or not el.query_id:
            continue
        raw = sdk.run_query(query_id=el.query_id, result_format="json")
        rows = json.loads(raw) if isinstance(raw, str) else raw
        name = el.title or f"Tile {el.id}"
        _write_sheet(wb, name, rows, title=name)
        exported.append({"tile": name, "rows": len(rows)})
    if not exported:
        return json.dumps({"error": "El dashboard no tiene tiles de datos exportables."})
    url = _upload_and_sign(_finish_workbook(wb),
                           _xlsx_name(filename or f"{db.title}_{dashboard_id}"))
    return json.dumps({"download_url": url, "dashboard": db.title,
                       "tiles": exported, "expires_in_hours": SIGNED_URL_HOURS},
                      ensure_ascii=False)


def _xlsx_name(name: str) -> str:
    stem = re.sub(r"[^\w\-]+", "_", name).strip("_") or "export"
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{stem}_{stamp}.xlsx"


ALL_TOOLS = [export_query_to_excel, export_multi_sheet_excel,
             export_dashboard_to_excel, list_excel_templates]
